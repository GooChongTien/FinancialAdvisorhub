#!/usr/bin/env python3
"""
Generate comprehensive navigation data map as Excel file
Includes all 217+ fields across 14 pages
"""

import pandas as pd
from datetime import datetime

# Define all columns
columns = [
    "Current Screen / Module",
    "Navigate From",
    "Navigate To",
    "Screen Section",
    "Section Field",
    "Description / Key Interactions",
    "Supabase Table",
    "Supabase Column Name",
    "Data Type",
    "Editable",
    "Validation Rule / Notes"
]

# Complete data - all 217+ rows
data = [
    # LOGIN & AUTH SECTION
    ["🔐 LOGIN & AUTH", "", "", "", "", "", "", "", "", "", ""],
    ["/login", "None, Logout, Unauthenticated access", "/register, / (home)", "Login Form", "—", "User authentication screen with email/password and forgot password flow.", "—", "—", "—", "—", "Entry point for all users"],
    ["/login", "(entry)", "/ (home), /register", "Login Form", "Email Address", "User email for authentication", "auth.users", "email", "string", "Y", "Required, valid email format"],
    ["/login", "(entry)", "/ (home), /register", "Login Form", "Password", "User password for authentication", "auth.users", "encrypted_password", "string", "Y", "Required, min 8 chars"],
    ["/login", "(entry)", "/ (home)", "Login Form", "Forgot Password Link", "Triggers password reset email", "auth", "—", "action", "Y", "Opens reset flow via Supabase Auth"],
    ["/login", "(recovery link)", "/ (home)", "Password Recovery", "New Password", "Set new password after reset", "auth.users", "encrypted_password", "string", "Y", "Min 8 chars, must match confirmation"],
    ["/login", "(recovery link)", "/ (home)", "Password Recovery", "Confirm New Password", "Password confirmation field", "—", "—", "string", "Y", "Must match New Password"],

    ["/register", "/login", "/login", "Registration Form", "—", "New advisor registration form, redirects to login on success.", "users", "—", "—", "—", "Admin may disable self-registration"],
    ["/register", "/login", "/login", "Registration Form", "Full Name", "Advisor's full name", "users", "full_name", "string", "Y", "Required, max 100 chars"],
    ["/register", "/login", "/login", "Registration Form", "Email Address", "Unique email for account", "users", "email", "string", "Y", "Required, valid email, must be unique"],
    ["/register", "/login", "/login", "Registration Form", "Mobile Number", "Contact phone number", "users", "phone", "string", "Y", "Required, regex: ^[0-9]{8,15}$"],
    ["/register", "/login", "/login", "Registration Form", "Password", "Account password", "auth.users", "encrypted_password", "string", "Y", "Min 8 chars, must contain uppercase + number"],
    ["/register", "/login", "/login", "Registration Form", "Confirm Password", "Password confirmation", "—", "—", "string", "Y", "Must match Password"],

    # HOME DASHBOARD SECTION
    ["🏠 HOME DASHBOARD", "", "", "", "", "", "", "", "", "", ""],
    ["/", "/login, /profile-settings, Sidebar", "/customers, /quick-quote, /todo, /analytics, /broadcast, /customers/detail?id={hotLeadId}", "Dashboard", "—", "Main advisor dashboard with widgets and quick actions", "—", "—", "—", "—", "Personalized content based on logged-in advisor"],
    ["/", "Sidebar", "(various)", "Header", "Greeting", "Personalized welcome message with advisor name", "users", "full_name", "string", "N", 'Format: "Good Morning, {name}"'],
    ["/", "Sidebar", "/customers", "Metrics", "Total Leads", "Count of all leads assigned to advisor", "leads", "COUNT(*)", "number", "N", "Calculated field"],
    ["/", "Sidebar", "/customers?filter=is_client", "Metrics", "Total Clients", "Count of leads where is_client=true", "leads", "COUNT(*) WHERE is_client=true", "number", "N", "Calculated field"],
    ["/", "Sidebar", "/todo?filter=pending", "Metrics", "Pending Tasks", "Count of incomplete tasks", "tasks", "COUNT(*) WHERE completed=false", "number", "N", "Calculated field"],
    ["/", "Sidebar", "/customers?filter=hot", "Hot Leads Widget", "Hot Lead Card", "Shows leads with high temperature score", "leads", "name, contact_number, last_contacted", "mixed", "N", "Click navigates to /customers/detail?id={id}"],
    ["/", "Sidebar", "/todo", "Reminders Widget", "Upcoming Task", "Shows next 3 tasks/appointments by date", "tasks", "title, date, time, type", "mixed", "N", "Click navigates to /todo"],
    ["/", "Sidebar", "/analytics", "Performance Widget", "RP/SP Progress", "Current month RP/SP vs targets", "analytics", "rp_incepted, rp_target, sp_incepted, sp_target", "number", "N", "Progress bars with percentages"],
    ["/", "Sidebar", "/broadcast", "Broadcast Feed Widget", "Latest Announcement", "Shows most recent broadcast with unread indicator", "broadcasts", "title, category, published_date", "mixed", "N", "Click navigates to /broadcast/detail?id={id}"],
    ["/", "Sidebar", "/quick-quote", "Quick Actions", "Quick Quote Button", "Opens quick quote calculator", "—", "—", "action", "—", "Fast path to quotation tool"],
    ["/", "Sidebar", "/customers/detail?id=new", "Quick Actions", "New Lead Button", "Opens new lead capture form", "—", "—", "action", "—", "Fast path to lead creation"],

    # CUSTOMER MANAGEMENT SECTION
    ["👥 CUSTOMER MANAGEMENT", "", "", "", "", "", "", "", "", "", ""],
    ["/customers", "Sidebar, / (home)", "/customers/detail?id={id}, /customers/detail?id=new", "Customer List", "—", "Lead and client list with search, filters, sorting, and new lead creation.", "leads", "—", "—", "—", "Main CRM hub"],
    ["/customers", "Sidebar", "/customers/detail?id=new", "Action Bar", "New Lead Button", "Creates new lead record", "leads", "—", "action", "Y", "Opens new lead form"],
    ["/customers", "(self)", "(self)", "Filters", "Search Input", "Free text search across name, email, contact", "leads", "name, email, contact_number", "string", "Y", "Min 2 chars, debounced search"],
    ["/customers", "(self)", "(self)", "Filters", "Status Filter", "Filter by lead status", "leads", "status", "enum", "Y", "Options: All / Not Initiated / Contacted / Proposal"],
    ["/customers", "(self)", "(self)", "Filters", "Source Filter", "Filter by lead source", "leads", "lead_source", "enum", "Y", "Options: All / Referral / Social Media / Walk-in / Cold Call / Website / Event / Other"],
    ["/customers", "(self)", "(self)", "Filters", "Type Filter", "Filter by client type", "leads", "is_client", "boolean", "Y", "Options: All Leads / Clients Only / Leads Only"],
    ["/customers", "(self)", "(self)", "Filters", "Temperature Filter", "Filter by engagement level", "—", "(calculated)", "enum", "Y", "Options: All / Hot / Warm / Cold (based on last_contacted)"],
    ["/customers", "(self)", "(self)", "Filters", "Date Range Filter", "Filter by last contacted date", "leads", "last_contacted", "date", "Y", "Date range picker"],
    ["/customers", "(self)", "(self)", "Table", "Name Column", "Lead/client full name", "leads", "name", "string", "N", "Click navigates to detail"],
    ["/customers", "(self)", "(self)", "Table", "Contact Column", "Phone number", "leads", "contact_number", "string", "N", "Click-to-call functionality"],
    ["/customers", "(self)", "(self)", "Table", "Email Column", "Email address", "leads", "email", "string", "N", "Click-to-email functionality"],
    ["/customers", "(self)", "(self)", "Table", "Status Badge", "Current lead status", "leads", "status", "enum", "N", "Color-coded: Not Initiated / Contacted / Proposal"],
    ["/customers", "(self)", "(self)", "Table", "Source Tag", "Lead source", "leads", "lead_source", "enum", "N", "Tag display"],
    ["/customers", "(self)", "(self)", "Table", "Last Contacted", "Last interaction date", "leads", "last_contacted", "date", "N", "Relative time format (e.g., '2 days ago')"],
    ["/customers", "(self)", "(self)", "Table", "Temperature Indicator", "Engagement level indicator", "—", "(calculated)", "visual", "N", "Hot (≤7 days) / Warm (≤30 days) / Cold (>30 days)"],
    ["/customers", "(self)", "(self)", "Table", "Action Menu", "Quick actions dropdown", "—", "—", "action", "Y", "Edit, Delete, Schedule Appointment"],

    # CUSTOMER DETAIL - OVERVIEW TAB
    ["/customers/detail", "/customers, / (home hot lead)", "/customers, /new-business?leadId={id}, /todo/new?leadId={id}", "Overview Tab", "—", "Customer detail with tabs: Overview, Portfolio, Servicing, Gap & Opportunity.", "leads", "—", "—", "—", "Full customer profile"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Full Name", "Customer full name", "leads", "name", "string", "Y", "Required, max 150 chars"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Contact Number", "Primary phone number", "leads", "contact_number", "string", "Y", "Required, regex: ^[0-9]{8,15}$"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Email Address", "Email address", "leads", "email", "string", "Y", "Valid email format"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "National ID", "NRIC or ID number", "leads", "national_id", "string", "Y", "Country-specific validation"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Date of Birth", "Birth date for age calculation", "leads", "date_of_birth", "date", "Y", "ISO date format, age must be 18-100"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Gender", "Gender selection", "leads", "gender", "enum", "Y", "Options: Male / Female / Other"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Marital Status", "Marital status selection", "leads", "marital_status", "enum", "Y", "Options: Single / Married / Divorced / Widowed"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Occupation", "Current occupation", "leads", "occupation", "string", "Y", "Free text, max 100 chars"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Address", "Residential address", "leads", "address", "string", "Y", "Free text, max 300 chars"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Nationality", "Nationality/citizenship", "leads", "nationality", "string", "Y", "Country dropdown"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Smoker Status", "Smoking status checkbox", "leads", "smoker_status", "boolean", "Y", "Affects premium calculations"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Lead Source", "How lead was acquired", "leads", "lead_source", "enum", "Y", "Options: Referral / Social Media / Walk-in / Cold Call / Website / Event / Other"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Lead Status", "Current pipeline stage", "leads", "status", "enum", "Y", "Options: Not Initiated / Contacted / Proposal"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Last Contacted", "Last interaction timestamp", "leads", "last_contacted", "datetime", "N", "Auto-updated on interactions"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Is Client", "Client conversion flag", "leads", "is_client", "boolean", "N", "Auto-set when first policy incepted"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Notes", "Free-form advisor notes", "leads", "notes", "text", "Y", "Rich text editor, no max length"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Age (Calculated)", "Calculated from date_of_birth", "—", "DATEDIFF(date_of_birth, NOW())", "number", "N", "Display only"],
    ["/customers/detail", "/customers", "/customers", "Overview Tab", "Temperature (Calculated)", "Engagement level based on last_contacted", "—", "(formula)", "visual", "N", "Hot/Warm/Cold indicator"],

    # CUSTOMER DETAIL - PORTFOLIO TAB
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "—", "List of active policies for this client", "policies", "—", "—", "—", "Empty if is_client=false"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Policy Number", "Unique policy identifier", "policies", "policy_number", "string", "N", "Click to view policy detail"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Product Name", "Insurance product name", "policies", "product_name", "string", "N", "—"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Coverage Type", "Type of coverage", "policies", "coverage_type", "enum", "N", "Hospitalisation / Death / CI / TPD / etc."],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Sum Assured", "Coverage amount", "policies", "sum_assured", "number", "N", "Currency format"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Premium Amount", "Premium per payment", "policies", "premium_amount", "number", "N", "Currency format"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Premium Frequency", "Payment frequency", "policies", "premium_frequency", "enum", "N", "Monthly / Quarterly / Semi-Annual / Annual"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Policy Status", "Current policy status", "policies", "status", "enum", "N", "Active / Lapsed / Surrendered / Matured"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Total Premium (Calculated)", "Sum of all active policy premiums", "—", "SUM(premium_amount) WHERE status='Active'", "number", "N", "Annualized total"],
    ["/customers/detail", "/customers", "/policies/detail?id={policyId}", "Portfolio Tab", "Total Coverage (Calculated)", "Sum of all sum_assured", "—", "SUM(sum_assured) WHERE status='Active'", "number", "N", "Total protection value"],

    # CUSTOMER DETAIL - SERVICING TAB
    ["/customers/detail", "/customers", "(modal)", "Servicing Tab", "—", "Shows available servicing options and past requests", "servicing_requests", "—", "—", "—", "Claims, renewals, policy changes"],
    ["/customers/detail", "/customers", "(modal)", "Servicing Tab", "Service Type", "Type of service request", "servicing_requests", "service_type", "enum", "Y", "Options: Claim / Renewal / Policy Change / Surrender"],
    ["/customers/detail", "/customers", "(modal)", "Servicing Tab", "Request Date", "Date of service request", "servicing_requests", "request_date", "datetime", "N", "Auto-set on creation"],
    ["/customers/detail", "/customers", "(modal)", "Servicing Tab", "Request Status", "Current request status", "servicing_requests", "status", "enum", "N", "Pending / In Progress / Completed / Rejected"],
    ["/customers/detail", "/customers", "(modal)", "Servicing Tab", "Request Details", "Description of service needed", "servicing_requests", "details", "text", "Y", "Free text, max 1000 chars"],

    # CUSTOMER DETAIL - GAP & OPPORTUNITY TAB
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "—", "Gap analysis comparing current vs recommended coverage", "gap_analysis", "—", "—", "—", "AI-generated recommendations"],
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "Coverage Type", "Type of coverage analyzed", "—", "(calculated)", "enum", "N", "Based on FNA recommendations"],
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "Current Coverage", "Existing sum assured", "—", "SUM(policies.sum_assured) GROUP BY coverage_type", "number", "N", "From active policies"],
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "Recommended Coverage", "Suggested sum assured", "—", "(from FNA calculation)", "number", "N", "Based on needs analysis"],
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "Gap Amount", "Shortfall in coverage", "—", "recommended - current", "number", "N", "Calculated field"],
    ["/customers/detail", "/customers", "(self)", "Gap & Opportunity Tab", "Gap Percentage", "Gap as percentage", "—", "(gap / recommended) * 100", "number", "N", "Percentage display"],
    ["/customers/detail", "/customers", "/new-business?leadId={id}", "Gap & Opportunity Tab", "Propose Solution Button", "Creates new proposal to address gap", "—", "—", "action", "Y", "Opens proposal workflow"],

    # NEW BUSINESS SECTION
    ["💼 NEW BUSINESS", "", "", "", "", "", "", "", "", "", ""],
    ["/new-business", "Sidebar, /customers/detail", "/proposals/detail?id={id}, /proposals/detail?id=new", "Proposal List", "—", "Proposal list with filters and smart redirect logic for new proposals.", "proposals", "—", "—", "—", "Main sales pipeline"],
    ["/new-business", "Sidebar", "/proposals/detail?id=new", "Action Bar", "New Proposal Button", "Creates new proposal record", "proposals", "—", "action", "Y", "Opens client selection modal"],
    ["/new-business", "(self)", "(self)", "Filters", "Search Input", "Search by proposal number or client name", "proposals", "proposal_number, proposer_name", "string", "Y", "Min 2 chars, debounced search"],
    ["/new-business", "(self)", "(self)", "Filters", "Stage Filter", "Filter by proposal stage", "proposals", "stage", "enum", "Y", "Options: All / Fact Finding / Financial Planning / Recommendation / Quotation / Application"],
    ["/new-business", "(self)", "(self)", "Filters", "Status Filter", "Filter by proposal status", "proposals", "status", "enum", "Y", "Options: All / In Progress / Pending for UW / Pending for Payment / Pending for Approval / Completed / Cancelled"],
    ["/new-business", "(self)", "(self)", "Filters", "Date Range Filter", "Filter by last updated date", "proposals", "last_updated", "date", "Y", "Date range picker"],
    ["/new-business", "(self)", "(self)", "Table", "Proposal Number", "Unique proposal ID", "proposals", "proposal_number", "string", "N", "Click navigates to detail"],
    ["/new-business", "(self)", "(self)", "Table", "Client Name", "Proposer name", "proposals", "proposer_name", "string", "N", "Link to customer detail"],
    ["/new-business", "(self)", "(self)", "Table", "Stage Badge", "Current workflow stage", "proposals", "stage", "enum", "N", "Color-coded badge"],
    ["/new-business", "(self)", "(self)", "Table", "Status Badge", "Current status", "proposals", "status", "enum", "N", "Color-coded badge"],
    ["/new-business", "(self)", "(self)", "Table", "Progress Bar", "Completion percentage", "proposals", "completion_percentage", "number", "N", "Visual progress bar (0-100%)"],
    ["/new-business", "(self)", "(self)", "Table", "Last Updated", "Last modification timestamp", "proposals", "last_updated", "datetime", "N", "Relative time format"],
    ["/new-business", "(self)", "(self)", "Table", "Action Menu", "Quick actions dropdown", "—", "—", "action", "Y", "Edit, Delete, Duplicate"],

    # PROPOSAL DETAIL - WORKFLOW STAGES
    ["/proposals/detail", "/new-business, /quote-summary", "/new-business", "Stage Progress", "—", "Full proposal workflow: Fact Find, FNA, Recommendation, Quotation, Application.", "proposals", "—", "—", "—", "Multi-stage wizard"],
    ["/proposals/detail", "/new-business", "/new-business", "Navigation", "Stage Tabs", "Horizontal stage indicator", "proposals", "stage", "enum", "Y", "Must follow stage order"],
    ["/proposals/detail", "/new-business", "/new-business", "Navigation", "Progress Percentage", "Overall completion", "proposals", "completion_percentage", "number", "N", "Auto-calculated based on filled fields"],

    # FACT FINDING STAGE
    ["/proposals/detail", "/new-business", "/new-business", "Fact Finding", "Personal Details Subsection", "Client personal information", "proposals", "fact_finding_data.personal", "object", "Y", "Name, DOB, NRIC, Gender, Marital Status, Occupation"],
    ["/proposals/detail", "/new-business", "/new-business", "Fact Finding", "Dependents Subsection", "Dependent information", "proposals", "fact_finding_data.dependents", "array", "Y", "Name, DOB, Relationship for each dependent"],
    ["/proposals/detail", "/new-business", "/new-business", "Fact Finding", "CKA Subsection", "Customer Knowledge Assessment", "proposals", "fact_finding_data.cka", "object", "Y", "Investment knowledge, risk tolerance questions"],
    ["/proposals/detail", "/new-business", "/new-business", "Fact Finding", "RPQ Subsection", "Risk Profile Questionnaire", "proposals", "fact_finding_data.rpq", "object", "Y", "5-10 risk assessment questions"],
    ["/proposals/detail", "/new-business", "/new-business", "Fact Finding", "Completion Formula", "Stage progress calculation", "—", "(personal 40% + dependents 20% + cka 20% + rpq 20%)", "number", "N", "Auto-calculated"],

    # FINANCIAL PLANNING STAGE
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Income Section", "Monthly income sources", "proposals", "fna_data.income", "array", "Y", "Source, Amount, Frequency for each income"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Expenses Section", "Monthly expenses", "proposals", "fna_data.expenses", "array", "Y", "Category, Amount for each expense"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Assets Section", "Current assets", "proposals", "fna_data.assets", "array", "Y", "Asset Type, Value for each asset"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Liabilities Section", "Current liabilities", "proposals", "fna_data.liabilities", "array", "Y", "Liability Type, Amount for each liability"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Existing Coverage Section", "Current insurance policies", "proposals", "fna_data.existing_coverage", "array", "Y", "Policy Type, Sum Assured, Premium for each"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Net Worth (Calculated)", "Assets minus liabilities", "—", "SUM(assets) - SUM(liabilities)", "number", "N", "Auto-calculated"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Affordability (Calculated)", "Monthly surplus", "—", "SUM(income) - SUM(expenses)", "number", "N", "Auto-calculated"],
    ["/proposals/detail", "/new-business", "/new-business", "Financial Planning", "Human Life Value (Calculated)", "Economic value calculation", "—", "(formula based on income, age, dependents)", "number", "N", "Auto-calculated"],

    # RECOMMENDATION STAGE
    ["/proposals/detail", "/new-business", "/new-business", "Recommendation", "Recommended Products", "AI-suggested products", "proposals", "recommendation_data.products", "array", "Y", "Product selection with customization"],
    ["/proposals/detail", "/new-business", "/new-business", "Recommendation", "Recommendation Rationale", "Explanation of recommendations", "proposals", "recommendation_data.rationale", "text", "Y", "Auto-generated, editable"],
    ["/proposals/detail", "/new-business", "/new-business", "Recommendation", "Client Confirmation", "Client acknowledgment", "proposals", "recommendation_data.client_confirmed", "boolean", "Y", "Required before proceeding to quotation"],
    ["/proposals/detail", "/new-business", "/new-business", "Recommendation", "Advisor Signature", "Digital signature capture", "proposals", "recommendation_data.advisor_signature", "string", "Y", "Base64 image data"],
    ["/proposals/detail", "/new-business", "/new-business", "Recommendation", "Documents Attached", "Supporting documents", "proposals", "recommendation_data.documents", "array", "Y", "File uploads (PDF, images)"],

    # QUOTATION STAGE
    ["/proposals/detail", "/new-business", "/new-business", "Quotation", "Life Assured Details", "Insured person information", "proposals", "quotation_data.life_assured", "object", "Y", "Can be different from proposer"],
    ["/proposals/detail", "/new-business", "/new-business", "Quotation", "Scenarios", "Multiple quote scenarios", "proposals", "quotation_data.scenarios", "array", "Y", "Product variations for comparison"],
    ["/proposals/detail", "/new-business", "/new-business", "Quotation", "Selected Products", "Final product selection", "proposals", "quotation_data.selected_products", "array", "Y", "Products to include in application"],
    ["/proposals/detail", "/new-business", "/new-business", "Quotation", "Premium Summary", "Total premium calculation", "proposals", "quotation_data.premium_summary", "object", "N", "Auto-calculated from selected products"],
    ["/proposals/detail", "/new-business", "/new-business", "Quotation", "Compare Quotes Button", "Side-by-side comparison", "—", "—", "action", "Y", "Opens comparison modal (max 5 scenarios)"],

    # APPLICATION STAGE
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Applicant Information", "Full application details", "proposals", "application_data.applicant", "object", "Y", "Complete personal and contact info"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Beneficiary Nomination", "Beneficiary details", "proposals", "application_data.beneficiaries", "array", "Y", "Name, NRIC, Relationship, Share %"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Underwriting Questions", "Medical and lifestyle questions", "proposals", "application_data.underwriting", "object", "Y", "15-20 yes/no questions + details"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Payment Details", "Payment method and authorization", "proposals", "application_data.payment", "object", "Y", "Bank account or credit card details"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Declarations", "Required declarations and consents", "proposals", "application_data.declarations", "object", "Y", "Multiple checkboxes (PDPA, T&C, etc.)"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Client Signature", "Final signature capture", "proposals", "application_data.client_signature", "string", "Y", "Base64 image data"],
    ["/proposals/detail", "/new-business", "/new-business", "Application", "Submit Button", "Final submission action", "—", "—", "action", "Y", "Generates PDF, submits to underwriting"],

    # QUICK QUOTE SECTION
    ["🧮 QUICK QUOTE", "", "", "", "", "", "", "", "", "", ""],
    ["/quick-quote", "Sidebar, / (home quick quote)", "/quote-summary", "Quick Quote Form", "—", "Fast quotation tool with product selection and premium calculation.", "quick_quotes", "—", "—", "—", "Instant quote calculator"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Product Selection", "Select insurance product", "quick_quotes", "product_id", "string", "Y", "Dropdown with product list"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Age", "Life assured age", "quick_quotes", "age", "number", "Y", "Range: 18-70"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Gender", "Life assured gender", "quick_quotes", "gender", "enum", "Y", "Options: Male / Female"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Smoker Status", "Smoking status", "quick_quotes", "smoker_status", "boolean", "Y", "Checkbox"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Sum Assured", "Coverage amount", "quick_quotes", "sum_assured", "number", "Y", "Min: 50000, currency input"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Premium Term", "Payment duration", "quick_quotes", "premium_term", "number", "Y", "Years, range: 5-30"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Coverage Term", "Coverage duration", "quick_quotes", "coverage_term", "number", "Y", "Years, range: 5-whole life"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Input Section", "Payment Frequency", "Payment frequency", "quick_quotes", "payment_frequency", "enum", "Y", "Options: Monthly / Quarterly / Annual"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Result Section", "Premium Amount (Calculated)", "Monthly/Annual premium", "quick_quotes", "premium_amount", "number", "N", "Auto-calculated on input change"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Result Section", "Total Premiums (Calculated)", "Total payment over term", "—", "premium_amount * term * frequency_multiplier", "number", "N", "Display only"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Action Bar", "Calculate Button", "Trigger calculation", "—", "—", "action", "Y", "Validates and calculates premium"],
    ["/quick-quote", "Sidebar", "/quote-summary", "Action Bar", "View Summary Button", "Navigate to summary", "—", "—", "action", "Y", "Opens /quote-summary with query params"],

    # QUOTE SUMMARY
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary Display", "—", "Shows quote details and allows conversion to full proposal.", "quick_quotes", "—", "—", "—", "Quote review screen"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary", "Product Name", "Selected product", "quick_quotes", "product_id → products.product_name", "string", "N", "Read-only"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary", "Life Assured Details", "Age, Gender, Smoker status", "quick_quotes", "age, gender, smoker_status", "mixed", "N", "Read-only"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary", "Coverage Details", "Sum assured, terms", "quick_quotes", "sum_assured, premium_term, coverage_term", "mixed", "N", "Read-only"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary", "Premium Breakdown", "Premium amount and frequency", "quick_quotes", "premium_amount, payment_frequency", "mixed", "N", "Read-only, with total"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Summary", "Premium Illustration Table", "Year-by-year breakdown", "—", "(calculated)", "table", "N", "Generated illustration"],
    ["/quote-summary", "/quick-quote", "/quick-quote, /proposals/detail?id=new", "Action Bar", "Back Button", "Return to calculator", "—", "—", "action", "Y", "Back to /quick-quote"],
    ["/quote-summary", "/quick-quote", "/proposals/detail?id=new", "Action Bar", "Convert to Proposal Button", "Create full proposal", "—", "—", "action", "Y", "Opens client selection, creates proposal with quote data"],
    ["/quote-summary", "/quick-quote", "(download)", "Action Bar", "Export PDF Button", "Download quote PDF", "—", "—", "action", "Y", "Generates PDF illustration"],

    # ANALYTICS SECTION
    ["📊 ANALYTICS", "", "", "", "", "", "", "", "", "", ""],
    ["/analytics", "Sidebar, / (home performance widget)", "(self)", "Dashboard", "—", "Performance dashboard with metrics, charts, goal tracking, and export.", "analytics", "—", "—", "—", "Advisor performance tracking"],
    ["/analytics", "Sidebar", "(self)", "Header", "Period Selector", "Select time period", "—", "(sessionStorage)", "enum", "Y", "Options: This Month / This Quarter / This Year / Custom Range"],
    ["/analytics", "Sidebar", "(self)", "Header", "Export Button", "Export to Excel", "—", "—", "action", "Y", "Downloads .xlsx with all metrics"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "RP Target", "Regular premium target", "analytics", "rp_target", "number", "N", "From agency/admin"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "RP Incepted", "Regular premium achieved", "analytics", "rp_incepted", "number", "N", "Sum of incepted policies"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "RP Achievement % (Calculated)", "RP achievement percentage", "—", "(rp_incepted / rp_target) * 100", "number", "N", "Progress indicator"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "SP Target", "Single premium target", "analytics", "sp_target", "number", "N", "From agency/admin"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "SP Incepted", "Single premium achieved", "analytics", "sp_incepted", "number", "N", "Sum of incepted policies"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "SP Achievement % (Calculated)", "SP achievement percentage", "—", "(sp_incepted / sp_target) * 100", "number", "N", "Progress indicator"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "Total Cases", "Number of cases submitted", "analytics", "total_cases", "number", "N", "Count of proposals with status != 'Cancelled'"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "Cases Incepted", "Number of incepted policies", "analytics", "cases_incepted", "number", "N", "Count of policies with status = 'Active'"],
    ["/analytics", "Sidebar", "(self)", "KPI Cards", "Conversion Rate % (Calculated)", "Case conversion rate", "—", "(cases_incepted / total_cases) * 100", "number", "N", "Percentage display"],
    ["/analytics", "Sidebar", "(self)", "Charts", "Production Trend", "Monthly production chart", "analytics", "production_by_month", "timeseries", "N", "Line/bar chart, RP vs SP"],
    ["/analytics", "Sidebar", "(self)", "Charts", "Product Mix", "Product distribution", "analytics", "production_by_product", "pie", "N", "Pie/donut chart"],
    ["/analytics", "Sidebar", "(self)", "Charts", "Pipeline Funnel", "Sales pipeline stages", "proposals", "COUNT(*) GROUP BY stage", "funnel", "N", "Funnel visualization"],
    ["/analytics", "Sidebar", "(self)", "Charts", "Commission Earned", "Commission tracking", "analytics", "commission_earned", "number", "N", "Sum of commissions"],
    ["/analytics", "Sidebar", "(self)", "Tables", "Top Products", "Best-performing products", "—", "(aggregated)", "table", "N", "Product name, count, total premium"],
    ["/analytics", "Sidebar", "(self)", "Tables", "Recent Cases", "Latest submissions", "proposals", "recent 10", "table", "N", "Proposal number, client, status, date"],
    ["/analytics", "Sidebar", "(self)", "Insights", "AI Insights", "Auto-generated insights", "—", "(AI-generated)", "text", "N", "Recommendations and alerts"],
    ["/analytics", "Sidebar", "(self)", "Insights", "Shortfall Alert", "Below-target warning", "—", "(calculated if achievement < 80%)", "alert", "N", "Shows gap and recommended actions"],
    ["/analytics", "Sidebar", "(self)", "Insights", "Hotspot Opportunities", "Lead engagement suggestions", "—", "(calculated from customer data)", "list", "N", "Leads needing follow-up"],

    # TO-DO & CALENDAR SECTION
    ["✅ TO-DO & CALENDAR", "", "", "", "", "", "", "", "", "", ""],
    ["/todo", "Sidebar, / (home reminders)", "/todo/new, (self)", "Calendar View", "—", "Calendar and task manager with list/calendar views, filters, .ics export.", "tasks", "—", "—", "—", "Task and appointment management"],
    ["/todo", "Sidebar", "(self)", "Header", "View Toggle", "Switch between list and calendar", "—", "(sessionStorage)", "enum", "Y", "Options: List View / Calendar View"],
    ["/todo", "Sidebar", "/todo/new", "Header", "New Task Button", "Create new task/appointment", "tasks", "—", "action", "Y", "Opens task creation modal"],
    ["/todo", "Sidebar", "(download)", "Header", "Export .ics Button", "Export to calendar app", "—", "—", "action", "Y", "Downloads iCalendar file"],
    ["/todo", "Sidebar", "(self)", "Filters", "Task Type Filter", "Filter by type", "tasks", "type", "enum", "Y", "Options: All / Task / Appointment"],
    ["/todo", "Sidebar", "(self)", "Filters", "Completion Filter", "Filter by completion", "tasks", "completed", "boolean", "Y", "Options: All / Pending / Completed"],
    ["/todo", "Sidebar", "(self)", "Filters", "Linked Lead Filter", "Filter by associated lead", "tasks", "linked_lead_id", "string", "Y", "Dropdown of leads"],
    ["/todo", "Sidebar", "(self)", "Filters", "Date Range Filter", "Filter by date range", "tasks", "date", "date", "Y", "Date range picker"],
    ["/todo", "Sidebar", "(self)", "List View", "Task/Appointment Card", "Individual task display", "tasks", "—", "mixed", "Y", "Click to expand details"],
    ["/todo", "Sidebar", "(self)", "List View", "Title", "Task title", "tasks", "title", "string", "Y", "Required, max 200 chars"],
    ["/todo", "Sidebar", "(self)", "List View", "Type Badge", "Task or Appointment", "tasks", "type", "enum", "N", "Color-coded badge"],
    ["/todo", "Sidebar", "(self)", "List View", "Date & Time", "Scheduled date and time", "tasks", "date, time", "datetime", "Y", "Date required, time optional"],
    ["/todo", "Sidebar", "(self)", "List View", "Duration", "Appointment duration", "tasks", "duration", "string", "Y", "Minutes, only for appointments"],
    ["/todo", "Sidebar", "(self)", "List View", "Linked Lead Name", "Associated customer", "tasks", "linked_lead_name", "string", "N", "Link to customer detail"],
    ["/todo", "Sidebar", "(self)", "List View", "Completion Checkbox", "Mark as complete", "tasks", "completed", "boolean", "Y", "Updates server and local state"],
    ["/todo", "Sidebar", "(self)", "List View", "Notes", "Task notes/description", "tasks", "notes", "text", "Y", "Free text, expandable"],
    ["/todo", "Sidebar", "(self)", "List View", "Action Menu", "Edit, Delete actions", "—", "—", "action", "Y", "Dropdown menu"],
    ["/todo", "Sidebar", "(self)", "Calendar View", "Calendar Grid", "Month/week calendar display", "tasks", "date", "calendar", "Y", "Drag-drop rescheduling"],
    ["/todo", "Sidebar", "(self)", "Calendar View", "Task Markers", "Visual task indicators on calendar", "tasks", "date, type", "visual", "N", "Color-coded dots/bars"],
    ["/todo", "Sidebar", "(self)", "Calendar View", "Birthday Indicators", "Auto-generated birthday reminders", "leads", "date_of_birth", "visual", "N", "Special marker for lead birthdays"],

    # TO-DO NEW FORM
    ["/todo/new", "/todo, /customers/detail", "/todo", "Task Form", "—", "Create new task or appointment", "tasks", "—", "—", "—", "Modal form"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Title", "Task title", "tasks", "title", "string", "Y", "Required, max 200 chars"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Type", "Task or Appointment", "tasks", "type", "enum", "Y", "Radio buttons: Task / Appointment"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Date", "Scheduled date", "tasks", "date", "date", "Y", "Required, date picker, must be >= today"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Time", "Scheduled time", "tasks", "time", "time", "N", "Time picker, required if type=Appointment"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Duration", "Appointment duration", "tasks", "duration", "number", "N", "Minutes, required if type=Appointment, default 30"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Linked Lead", "Associate with customer", "tasks", "linked_lead_id, linked_lead_name", "string", "N", "Searchable dropdown"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Notes", "Task description", "tasks", "notes", "text", "Y", "Textarea, max 1000 chars"],
    ["/todo/new", "/todo", "/todo", "Task Form", "Save Button", "Create task", "—", "—", "action", "Y", "Validates and saves to DB"],

    # BROADCAST SECTION
    ["📻 BROADCAST", "", "", "", "", "", "", "", "", "", ""],
    ["/broadcast", "Sidebar, / (home broadcast)", "/broadcast/detail?id={id}", "Broadcast List", "—", "Announcements with categories, pinned posts, and unread indicators.", "broadcasts", "—", "—", "—", "Company announcements hub"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "Header", "Search Input", "Search announcements", "broadcasts", "title, content", "string", "Y", "Full-text search"],
    ["/broadcast", "Sidebar", "(self)", "Header", "Category Filter", "Filter by category", "broadcasts", "category", "enum", "Y", "Options: All / Announcement / Training / Campaign"],
    ["/broadcast", "Sidebar", "(self)", "Pinned Section", "Pinned Announcements", "Important pinned posts", "broadcasts", "WHERE is_pinned=true", "list", "N", "Always shown at top"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "Pinned Section", "Pinned Card", "Individual pinned announcement", "broadcasts", "—", "mixed", "N", "Title, category, date, unread indicator"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "Recent Section", "Recent Announcements", "Non-pinned posts by date", "broadcasts", "ORDER BY published_date DESC", "list", "N", "Chronological list"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "Recent Section", "Announcement Card", "Individual announcement", "broadcasts", "—", "mixed", "N", "Title, category, date, unread indicator"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "List Item", "Title", "Announcement title", "broadcasts", "title", "string", "N", "Click to view detail"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "List Item", "Category Badge", "Category tag", "broadcasts", "category", "enum", "N", "Color-coded: Announcement / Training / Campaign"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "List Item", "Published Date", "Publication timestamp", "broadcasts", "published_date", "datetime", "N", "Relative time format"],
    ["/broadcast", "Sidebar", "/broadcast/detail?id={id}", "List Item", "Unread Indicator", "Visual unread marker", "—", "(sessionStorage)", "visual", "N", "Yellow dot if not read in session"],

    # BROADCAST DETAIL
    ["/broadcast/detail", "/broadcast", "/broadcast", "Broadcast Detail", "—", "Full broadcast content view with attachments and auto-mark as read.", "broadcasts", "—", "—", "—", "Announcement detail screen"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Content", "Title", "Announcement title", "broadcasts", "title", "string", "N", "Read-only"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Content", "Category Badge", "Category tag", "broadcasts", "category", "enum", "N", "Read-only"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Content", "Published Date", "Publication timestamp", "broadcasts", "published_date", "datetime", "N", "Read-only"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Content", "Content Body", "Full announcement text", "broadcasts", "content", "text", "N", "Rich text display, read-only"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Content", "Attachments", "Attached files", "broadcasts", "attachments", "array", "N", "Download links for PDFs, images, etc."],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Action", "Mark as Read (Auto)", "Auto-mark on view", "—", "(sessionStorage)", "action", "N", "Triggers on page load"],
    ["/broadcast/detail", "/broadcast", "/broadcast", "Action", "Back Button", "Return to list", "—", "—", "action", "Y", "Navigate to /broadcast"],

    # POLICIES SECTION
    ["📄 POLICIES", "", "", "", "", "", "", "", "", "", ""],
    ["/policies/detail", "/customers/detail (portfolio)", "(back)", "Policy Detail", "—", "Policy details view with product info, coverage, beneficiaries, documents, and claims.", "policies", "—", "—", "—", "Read-only policy information"],
    ["/policies/detail", "/customers/detail", "(back)", "Header", "Policy Number", "Unique policy identifier", "policies", "policy_number", "string", "N", "Read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Header", "Policy Status Badge", "Current status", "policies", "status", "enum", "N", "Color-coded: Active / Lapsed / Surrendered / Matured"],
    ["/policies/detail", "/customers/detail", "(back)", "Basic Information", "Client Name", "Policyholder name", "policies", "client_name", "string", "N", "Read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Basic Information", "Product Name", "Insurance product", "policies", "product_name", "string", "N", "Read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Basic Information", "Coverage Type", "Type of coverage", "policies", "coverage_type", "enum", "N", "Read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Coverage Details", "Sum Assured", "Coverage amount", "policies", "sum_assured", "number", "N", "Currency format, read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Coverage Details", "Premium Amount", "Premium per payment", "policies", "premium_amount", "number", "N", "Currency format, read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Coverage Details", "Premium Frequency", "Payment frequency", "policies", "premium_frequency", "enum", "N", "Read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Coverage Details", "Annual Premium (Calculated)", "Annualized premium", "—", "premium_amount * frequency_multiplier", "number", "N", "Display only"],
    ["/policies/detail", "/customers/detail", "(back)", "Policy Dates", "Policy Start Date", "Inception date", "policies", "policy_start_date", "date", "N", "ISO date format, read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Policy Dates", "Policy End Date", "Maturity/expiry date", "policies", "policy_end_date", "date", "N", "ISO date format, read-only"],
    ["/policies/detail", "/customers/detail", "(back)", "Policy Dates", "Years Remaining (Calculated)", "Time to maturity", "—", "DATEDIFF(policy_end_date, NOW()) / 365", "number", "N", "Display only"],
    ["/policies/detail", "/customers/detail", "(back)", "Beneficiaries", "Beneficiary List", "Nominated beneficiaries", "policy_beneficiaries", "—", "table", "N", "Name, NRIC, Relationship, Share %"],
    ["/policies/detail", "/customers/detail", "(back)", "Documents", "Policy Document", "Official policy document", "policy_documents", "WHERE type='policy'", "file", "N", "PDF download link"],
    ["/policies/detail", "/customers/detail", "(back)", "Documents", "Benefit Illustration", "Benefit illustration document", "policy_documents", "WHERE type='illustration'", "file", "N", "PDF download link"],
    ["/policies/detail", "/customers/detail", "(back)", "Documents", "Premium Notices", "Premium payment notices", "policy_documents", "WHERE type='notice'", "list", "N", "Multiple PDFs, sorted by date"],
    ["/policies/detail", "/customers/detail", "(back)", "Claims History", "Claims List", "Past claims on this policy", "policy_claims", "—", "table", "N", "Claim date, type, amount, status"],
    ["/policies/detail", "/customers/detail", "(back)", "Action Bar", "Back Button", "Return to customer portfolio", "—", "—", "action", "Y", "Navigate back"],

    # PROFILE SETTINGS SECTION
    ["⚙️ PROFILE SETTINGS", "", "", "", "", "", "", "", "", "", ""],
    ["/profile-settings", "User dropdown", "/", "Profile Settings", "—", "Advisor profile, security, 2FA, preferences, back to home on exit.", "users", "—", "—", "—", "User account management"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Full Name", "Display and edit advisor full name", "users", "full_name", "string", "Y", "Max 100 chars"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Email Address", "Read-only email address", "users", "email", "string", "N", "Must be valid email format"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Mobile Number", "Editable mobile phone number", "users", "phone", "string", "Y", "Regex: ^[0-9]{8,15}$"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Advisor ID", "Unique internal advisor identifier", "users", "advisor_id", "string", "N", "Generated by admin"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Advisor ID Expiry", "Read-only expiry date (admin managed)", "users", "advisor_id_expiry", "date", "N", "ISO date format"],
    ["/profile-settings", "User dropdown", "/", "Personal Information", "Account Status", "Active/Inactive status (admin managed)", "users", "status", "enum", "N", "Values: Active / Inactive"],
    ["/profile-settings", "User dropdown", "/", "Security", "Two-Factor Authentication", "Toggle to enable/disable 2FA", "user_security", "two_factor_enabled", "boolean", "Y", "Default = false"],
    ["/profile-settings", "User dropdown", "/", "Security", "Change Password", "Opens change-password dialog (handled by Supabase Auth)", "auth", "password", "string", "Y", "Min length 8; handled via Supabase auth API"],
    ["/profile-settings", "User dropdown", "/", "User Preferences", "Language", "Dropdown selection for UI language", "user_preferences", "language", "enum", "Y", "Options: EN / ZH / MY / TH"],
    ["/profile-settings", "User dropdown", "/", "User Preferences", "Currency", "Dropdown selection for currency formatting", "user_preferences", "currency", "enum", "Y", "Options: SGD / MYR / USD"],
    ["/profile-settings", "User dropdown", "/", "Action Bar", "Save Changes Button", "Persist profile updates", "—", "—", "action", "Y", "Validates and saves to users table"],
    ["/profile-settings", "User dropdown", "/", "Action Bar", "Cancel Button", "Discard changes", "—", "—", "action", "Y", "Resets form to original values"],
]

# Create DataFrame
df = pd.DataFrame(data, columns=columns)

# Generate Excel file with formatting
output_file = "advisorhub-navigation-map-complete.xlsx"

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Navigation Map', index=False)

    # Get workbook and worksheet for formatting
    workbook = writer.book
    worksheet = writer.sheets['Navigation Map']

    # Set column widths
    worksheet.column_dimensions['A'].width = 25  # Current Screen
    worksheet.column_dimensions['B'].width = 35  # Navigate From
    worksheet.column_dimensions['C'].width = 35  # Navigate To
    worksheet.column_dimensions['D'].width = 25  # Screen Section
    worksheet.column_dimensions['E'].width = 25  # Section Field
    worksheet.column_dimensions['F'].width = 50  # Description
    worksheet.column_dimensions['G'].width = 20  # Supabase Table
    worksheet.column_dimensions['H'].width = 25  # Supabase Column
    worksheet.column_dimensions['I'].width = 12  # Data Type
    worksheet.column_dimensions['J'].width = 10  # Editable
    worksheet.column_dimensions['K'].width = 50  # Validation Rule

    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Format section headers (rows with emoji)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=12)

    for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1):
        if row[0].value and any(char in str(row[0].value) for char in ['🔐', '🏠', '👥', '💼', '🧮', '📊', '✅', '📻', '📄', '⚙️']):
            for cell in row:
                cell.fill = section_fill
                cell.font = section_font

    # Freeze header row
    worksheet.freeze_panes = 'A2'

    # Add autofilter
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"✅ Excel file generated successfully!")
print(f"📄 File: {output_file}")
print(f"📊 Total rows: {len(data)} (including {len([r for r in data if any(char in str(r[0]) for char in ['🔐', '🏠', '👥', '💼', '🧮', '📊', '✅', '📻', '📄', '⚙️'])])} section headers)")
print(f"📋 Columns: {len(columns)}")
print(f"💡 Features: Formatted headers, section highlights, frozen panes, auto-filter, optimized column widths")
